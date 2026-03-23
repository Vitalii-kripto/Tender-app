import uvicorn
import asyncio
import sys
import os
import logging
import tempfile
import openpyxl
import re
import sqlite3
from dotenv import load_dotenv

# Загружаем переменные окружения в самом начале
load_dotenv()
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell
from typing import List, Dict, Any

from fastapi import FastAPI, Depends, HTTPException, UploadFile, File, Body, BackgroundTasks
from fastapi.responses import FileResponse
from fastapi.middleware.cors import CORSMiddleware
from sqlalchemy.orm import Session

if sys.platform == 'win32':
    asyncio.set_event_loop_policy(asyncio.WindowsProactorEventLoopPolicy())

from .database import engine, Base, get_db
from .models import TenderModel, ProductModel
from .services.eis_service import EisService, Notice, mark_seen, csv_append_row, OUT_DIR
from .services.parser import GidroizolParser
from .services.document_service import DocumentService
from .services.ai_service import AiService
from .services.legal_analysis_service import LegalAnalysisService
from .services.batch_analysis import analyze_tenders_batch_job
from .services.job_service import job_service

def clean_markdown(text):
    """Удаляет markdown-артефакты из текста"""
    if not text: return ""
    # Удаляем жирный/курсив
    text = re.sub(r'\*\*|\*|__|_', '', text)
    # Удаляем заголовки
    text = re.sub(r'#+\s+', '', text)
    # Удаляем горизонтальные линии
    text = re.sub(r'^-{3,}\s*$', '', text, flags=re.MULTILINE)
    # Удаляем лишние пустые строки
    text = re.sub(r'\n{3,}', '\n\n', text)
    return text.strip()

def parse_markdown_table(text):
    """Парсит markdown-таблицу в список списков"""
    if not text or '|' not in text: return []
    lines = text.strip().split('\n')
    table_rows = []
    for line in lines:
        if '|' in line:
            # Пропускаем разделители |---|---|
            if re.match(r'^[\s|:\-\+]+$', line.strip()):
                continue
            # Разбиваем по |
            parts = [p.strip() for p in line.split('|')]
            # Убираем пустые элементы по краям, если строка начиналась/заканчивалась на |
            if line.strip().startswith('|'):
                parts = parts[1:]
            if line.strip().endswith('|'):
                parts = parts[:-1]
            
            if parts and any(p for p in parts):
                table_rows.append([clean_markdown(p) for p in parts])
    return table_rows

def parse_markdown_list(text):
    """Парсит markdown-список в список строк"""
    if not text: return []
    lines = text.strip().split('\n')
    list_items = []
    for line in lines:
        # Ищем маркеры списка: -, *, 1., 1)
        match = re.match(r'^\s*(?:[\-\*\+]|\d+[\.\)])\s+(.*)', line)
        if match:
            list_items.append(clean_markdown(match.group(1)))
        elif line.strip() and not re.match(r'^[\s|:\-\+]+$', line.strip()) and '|' not in line:
            # Если это просто строка текста, тоже берем ее как элемент, если она не пустая
            list_items.append(clean_markdown(line))
    return list_items

# --- LOGGING SETUP ---
log_file = "fastapi_app_log.txt"

# Diagnostic logs for encoding
print(f"DEBUG: sys.getdefaultencoding() = {sys.getdefaultencoding()}")
import locale
print(f"DEBUG: locale.getpreferredencoding(False) = {locale.getpreferredencoding(False)}")
print(f"DEBUG: os.environ.get('PYTHONUTF8') = {os.environ.get('PYTHONUTF8')}")

# Test case for Russian text
test_file = "russian_test.txt"
test_text = "Проверка кириллицы: Привет, мир!"
try:
    with open(test_file, "w", encoding="utf-8") as f:
        f.write(test_text)
    with open(test_file, "r", encoding="utf-8") as f:
        read_text = f.read()
    if read_text == test_text:
        print(f"✅ Russian text test passed: {read_text}")
    else:
        print(f"❌ Russian text test failed: expected '{test_text}', got '{read_text}'")
except Exception as e:
    print(f"❌ Russian text test error: {e}")

# Удаляем существующие хендлеры, если они есть, чтобы избежать дублирования при релоаде
for handler in logging.root.handlers[:]:
    logging.root.removeHandler(handler)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(log_file, encoding='utf-8', mode='w'), # mode='w' перезаписывает файл
        logging.StreamHandler()
    ]
)
logger = logging.getLogger("FastAPI_Main")

# --- SETUP ---
def migrate_db():
    """Простейшая миграция для добавления недостающих колонок в SQLite"""
    from .database import DB_PATH
    
    logger.info(f"Checking schema for {DB_PATH}...")
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    
    # Список колонок, которые могли быть добавлены позже
    # (column_name, table_name, column_type)
    new_columns = [
        ("docs_url", "tenders", "TEXT"),
        ("search_url", "tenders", "TEXT"),
        ("keyword", "tenders", "TEXT"),
        ("ntype", "tenders", "TEXT"),
        ("local_file_path", "tenders", "TEXT"),
        ("extracted_text", "tenders", "TEXT"),
        ("created_at", "tenders", "DATETIME"),
        ("description", "products", "TEXT"),
        ("updated_at", "products", "DATETIME"),
    ]
    
    for col_name, table_name, col_type in new_columns:
        try:
            cursor.execute(f"ALTER TABLE {table_name} ADD COLUMN {col_name} {col_type}")
            logger.info(f"Added column {col_name} to table {table_name}")
        except sqlite3.OperationalError:
            # Колонка уже существует
            pass
            
    conn.commit()
    conn.close()

try:
    logger.info("Initializing Database...")
    Base.metadata.create_all(bind=engine)
    migrate_db()
    logger.info("Database initialized successfully.")
except Exception as e:
    logger.critical(f"Database initialization failed: {e}", exc_info=True)

app = FastAPI(title="TenderSmart Gidroizol API", version="2.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # Разрешить все для локальной разработки
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Services
try:
    logger.info("Initializing Services...")
    eis_service = EisService()
    parser_service = GidroizolParser()
    doc_service = DocumentService()
    ai_service = AiService()
    legal_analysis_service = LegalAnalysisService(ai_service.client)
    logger.info("Services initialized.")
except Exception as e:
    logger.error(f"Service initialization error: {e}", exc_info=True)

# --- ENDPOINTS ---

@app.get("/")
def read_root():
    logger.info("Health check endpoint hit.")
    return {"status": "online", "system": "TenderSmart PRO Backend"}

# --- CRM ENDPOINTS (Database Sync) ---

@app.get("/api/crm/tenders")
def get_crm_tenders(db: Session = Depends(get_db)):
    """Получить все тендеры из базы"""
    logger.info("Fetching all CRM tenders.")
    return db.query(TenderModel).all()

@app.post("/api/crm/tenders")
def add_update_tender(background_tasks: BackgroundTasks, tender: dict = Body(...), db: Session = Depends(get_db)):
    """Добавить или обновить тендер в CRM"""
    logger.info(f"Add/Update tender request: {tender.get('id')}")
    try:
        existing = db.query(TenderModel).filter(TenderModel.id == tender['id']).first()
        
        # Parse initial_price to float
        raw_price = tender.get('initial_price', 0)
        parsed_price = 0.0
        if isinstance(raw_price, str):
            cleaned = re.sub(r'[^\d,.-]', '', raw_price).replace(',', '.')
            try:
                parsed_price = float(cleaned)
            except ValueError:
                parsed_price = 0.0
        else:
            parsed_price = float(raw_price)

        if existing:
            existing.status = tender.get('status', existing.status)
            existing.risk_level = tender.get('risk_level', existing.risk_level)
            logger.info(f"Updated existing tender: {tender['id']}")
        else:
            new_tender = TenderModel(
                id=tender['id'],
                title=tender['title'],
                description=tender.get('description', ''),
                initial_price=parsed_price,
                deadline=tender.get('deadline', '-'),
                status=tender.get('status', 'Found'),
                risk_level=tender.get('risk_level', 'Low'),
                region=tender.get('region', 'РФ'),
                law_type=tender.get('law_type', '44-ФЗ'),
                url=tender.get('url', ''),
                docs_url=tender.get('docs_url', ''),
                search_url=tender.get('search_url', ''),
                keyword=tender.get('keyword', ''),
                ntype=tender.get('ntype', '')
            )
            db.add(new_tender)
            logger.info(f"Created new tender: {tender['id']}")
            
            # Если это новый тендер, запускаем скачивание документов
            if tender.get('docs_url'):
                notice = Notice(
                    reg=tender['id'],
                    ntype=tender.get('ntype', ''),
                    keyword=tender.get('keyword', ''),
                    search_url=tender.get('search_url', ''),
                    href=tender.get('url', ''),
                    docs_url=tender.get('docs_url', ''),
                    title=tender.get('title', ''),
                    object_info=tender.get('description', ''),
                    initial_price=str(tender.get('initial_price', '')),
                    application_deadline=tender.get('deadline', '')
                )
                background_tasks.add_task(eis_service.process_tenders, [notice])
        
        db.commit()
        return {"status": "success"}
    except Exception as e:
        logger.error(f"Error saving tender: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/api/crm/tenders/{tender_id}")
def delete_tender(tender_id: str, db: Session = Depends(get_db)):
    """Удалить тендер из базы"""
    logger.info(f"Deleting tender: {tender_id}")
    tender = db.query(TenderModel).filter(TenderModel.id == tender_id).first()
    if tender:
        db.delete(tender)
        db.commit()
    return {"status": "deleted"}

# --- SEARCH & PARSING ---

@app.post("/api/search-tenders/cancel")
def cancel_search():
    """Отменить текущий поиск"""
    logger.info("Cancel search request received")
    eis_service.cancel_search()
    return {"status": "cancelled"}

@app.get("/api/search-tenders")
def search_tenders_endpoint(
    query: str, 
    fz44: bool = True, 
    fz223: bool = True, 
    only_application_stage: bool = True, 
    publish_days_back: int = 30
):
    """Поиск через Playwright"""
    logger.info(f"Search request received: {query}")
    try:
        notices = eis_service.search_tenders(
            query=query, 
            fz44=fz44, 
            fz223=fz223, 
            only_application_stage=only_application_stage, 
            publish_days_back=publish_days_back
        )
    except RuntimeError as e:
        logger.error(f"Search failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))
    except Exception as e:
        logger.error(f"Search failed with unexpected error: {e}")
        raise HTTPException(status_code=500, detail=f"Search failed: {str(e)}")
    
    # Convert Notice dataclass to dict for JSON response
    result = []
    for n in notices:
        raw_price = n.initial_price
        parsed_price = 0.0
        if isinstance(raw_price, str):
            cleaned = re.sub(r'[^\d,.-]', '', raw_price).replace(',', '.')
            try:
                parsed_price = float(cleaned)
            except ValueError:
                parsed_price = 0.0
        else:
            parsed_price = float(raw_price)

        result.append({
            "id": n.reg,
            "eis_number": n.reg,
            "title": n.title,
            "description": n.object_info,
            "initial_price": parsed_price,
            "initial_price_text": str(raw_price),
            "initial_price_value": parsed_price,
            "deadline": n.application_deadline,
            "status": "Found",
            "risk_level": "Low",
            "region": "РФ",
            "law_type": n.ntype,
            "url": n.href,
            "docs_url": f"https://zakupki.gov.ru/epz/order/notice/{n.ntype}/view/documents.html?regNumber={n.reg}",
            "search_url": n.search_url,
            "keyword": n.keyword,
            "ntype": n.ntype,
            "seen": n.seen
        })
    return result

@app.post("/api/search-tenders/process")
def process_tenders(background_tasks: BackgroundTasks, tenders: list = Body(...), db: Session = Depends(get_db)):
    """Обработать выбранные тендеры"""
    logger.info(f"Processing {len(tenders)} selected tenders")
    try:
        for tender in tenders:
            existing = db.query(TenderModel).filter(TenderModel.id == tender['id']).first()
            
            # Parse initial_price to float
            raw_price = tender.get('initial_price', 0)
            parsed_price = 0.0
            if isinstance(raw_price, str):
                cleaned = re.sub(r'[^\d,.-]', '', raw_price).replace(',', '.')
                try:
                    parsed_price = float(cleaned)
                except ValueError:
                    parsed_price = 0.0
            else:
                parsed_price = float(raw_price)

            if existing:
                existing.status = tender.get('status', existing.status)
                existing.risk_level = tender.get('risk_level', existing.risk_level)
                logger.info(f"Updated existing tender: {tender['id']}")
            else:
                new_tender = TenderModel(
                    id=tender['id'],
                    title=tender['title'],
                    description=tender.get('description', ''),
                    initial_price=parsed_price,
                    deadline=tender.get('deadline', '-'),
                    status=tender.get('status', 'Found'),
                    risk_level=tender.get('risk_level', 'Low'),
                    region=tender.get('region', 'РФ'),
                    law_type=tender.get('law_type', '44-ФЗ'),
                    url=tender.get('url', ''),
                    docs_url=tender.get('docs_url', ''),
                    search_url=tender.get('search_url', ''),
                    keyword=tender.get('keyword', ''),
                    ntype=tender.get('ntype', '')
                )
                db.add(new_tender)
                logger.info(f"Created new tender: {tender['id']}")
                
                # Если это новый тендер, запускаем скачивание документов
                if tender.get('docs_url'):
                    notice = Notice(
                        reg=tender['id'],
                        ntype=tender.get('ntype', ''),
                        keyword=tender.get('keyword', ''),
                        search_url=tender.get('search_url', ''),
                        href=tender.get('url', ''),
                        docs_url=tender.get('docs_url', ''),
                        title=tender.get('title', ''),
                        object_info=tender.get('description', ''),
                        initial_price=str(tender.get('initial_price', '')),
                        application_deadline=tender.get('deadline', '')
                    )
                    background_tasks.add_task(eis_service.process_tenders, [notice])
            
        db.commit()
        return {"status": "success", "processed": len(tenders)}
    except Exception as e:
        logger.error(f"Error processing tenders: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/search-tenders/skip")
def skip_tender(tender: dict = Body(...)):
    """Пропустить тендер (отметить как просмотренный)"""
    logger.info(f"Skipping tender: {tender.get('id')}")
    try:
        mark_seen(tender['id'])
        csv_append_row(
            tender['id'], 
            tender.get('ntype', ''), 
            tender.get('keyword', ''), 
            tender.get('search_url', ''), 
            tender.get('docs_url', ''), 
            "SKIP:user_not_selected", 
            tender.get('title', ''), 
            tender.get('description', ''), 
            str(tender.get('initial_price', '')), 
            tender.get('deadline', '')
        )
        return {"status": "success"}
    except Exception as e:
        logger.error(f"Error skipping tender: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/products")
def get_products_endpoint(db: Session = Depends(get_db)):
    """Получение сохраненных товаров из БД без запуска парсера"""
    logger.info("Fetching products from DB.")
    products = db.query(ProductModel).all()
    result = []
    for p in products:
        result.append({
            "id": str(p.id),
            "title": p.title,
            "category": p.category,
            "material_type": p.material_type,
            "price": p.price,
            "specs": p.specs if p.specs else {},
            "url": p.url,
            "description": p.description # Added description
        })
    return result

@app.get("/api/parse-catalog")
async def parse_catalog_endpoint(db: Session = Depends(get_db)):
    """Запуск парсера каталога Gidroizol.ru и обновление БД"""
    logger.info("Starting catalog parser manually.")
    
    # Новый парсер возвращает количество сохраненных записей (int), а не список объектов
    saved_count = await parser_service.parse_and_save(db)
    logger.info(f"Parser finished. Saved/Updated {saved_count} items.")
    
    # После парсинга забираем актуальные данные из БД
    products = db.query(ProductModel).all()
    
    result = []
    for p in products:
        result.append({
            "id": str(p.id),
            "title": p.title,
            "category": p.category,
            "material_type": p.material_type,
            "price": p.price,
            "specs": p.specs if p.specs else {},
            "url": p.url,
            "description": p.description # Added description
        })
    return result

# --- AI & DOCS ENDPOINTS ---

@app.get("/api/tenders/{tender_id}/files")
def get_tender_files(tender_id: str):
    """Получить список скачанных файлов для тендера"""
    logger.info(f"Fetching files for tender {tender_id}")
    tender_dir = os.path.join(OUT_DIR, tender_id)
    if not os.path.exists(tender_dir):
        return []
    
    files = []
    for filename in os.listdir(tender_dir):
        filepath = os.path.join(tender_dir, filename)
        if os.path.isfile(filepath):
            files.append({
                "name": filename,
                "size": os.path.getsize(filepath),
                "ext": os.path.splitext(filename)[1].lower()
            })
    return files

@app.post("/api/ai/analyze-tenders-batch")
async def api_analyze_tenders_batch(background_tasks: BackgroundTasks, data: dict = Body(...)):
    logger.info("Batch AI Analysis request received.")
    tender_ids = data.get('tender_ids', [])
    selected_files = data.get('selected_files', {}) # {tender_id: [filenames]}
    
    if not tender_ids:
        raise HTTPException(status_code=400, detail="No tender IDs provided")
    
    job_id = job_service.create_job(tender_ids)
    background_tasks.add_task(analyze_tenders_batch_job, job_id, tender_ids, doc_service, legal_analysis_service, selected_files)
    
    return {"job_id": job_id}

@app.get("/api/ai/jobs/{job_id}")
async def get_job_status(job_id: str):
    job = job_service.get_job(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    return job

@app.post("/api/ai/extract-details")
async def api_extract_details(data: dict = Body(...)):
    """Извлечение данных о тендере из текста"""
    logger.info("AI Extract Details request.")
    text = data.get('text', '')
    return ai_service.extract_tender_details(text)

@app.post("/api/ai/extract-products")
async def api_extract_products(data: dict = Body(...)):
    """Извлечение списка товаров из сметы/КП"""
    logger.info("AI Extract Products request.")
    text = data.get('text', '')
    return ai_service.extract_products_from_text(text)

@app.post("/api/ai/enrich-specs")
async def api_enrich_specs(data: dict = Body(...)):
    """Поиск характеристик товара в интернете"""
    logger.info("AI Enrich Specs request.")
    product_name = data.get('product_name', '')
    result = ai_service.enrich_product_specs(product_name)
    return {"specs": result}

@app.post("/api/ai/match-product")
async def api_match_product(data: dict = Body(...), db: Session = Depends(get_db)):
    specs = data.get('specs', '')
    mode = data.get('mode', 'database') # 'database' or 'internet'
    logger.info(f"AI Match Product request. Mode: {mode}, Query len: {len(specs)}")

    if mode == 'internet':
        # Поиск в интернете через Grounding
        result_text = ai_service.search_products_internet(specs)
        return {"mode": "internet", "text": result_text}
    else:
        # Поиск по базе
        products_db = db.query(ProductModel).limit(50).all()
        catalog = [{"id": str(p.id), "title": p.title, "specs": p.specs} for p in products_db]
        matches = ai_service.find_product_equivalent(specs, catalog)
        return {"mode": "database", "matches": matches}

@app.post("/api/ai/validate-compliance")
async def api_validate_compliance(data: dict = Body(...)):
    """Валидация ТЗ vs Материал (Complex)"""
    logger.info("AI Compliance Validation request.")
    requirements = data.get('requirements', '')
    proposal = data.get('proposal', '[]')
    return ai_service.compare_requirements_vs_proposal(requirements, proposal)

@app.post("/api/ai/check-compliance")
async def api_check_compliance(data: dict = Body(...)):
    """Проверка пакета документов"""
    logger.info("AI Document Package Check request.")
    return ai_service.check_compliance(data['title'], data['description'], data['filenames'])

@app.post("/api/tenders/upload")
async def upload_file(file: UploadFile = File(...)):
    logger.info(f"File upload request: {file.filename}")
    try:
        file_path = await doc_service.save_file(file)
        # Запускаем OCR или извлечение текста
        text = doc_service.extract_text(file_path)
        logger.info("File processed successfully.")
        return {"text": text, "path": file_path}
    except Exception as e:
        logger.error(f"Upload Error: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/dashboard-stats")
async def get_dashboard_stats(db: Session = Depends(get_db)):
    logger.info("Dashboard stats requested.")
    count = db.query(TenderModel).count()
    return {
        "active_tenders": count,
        "margin_val": "₽14.2M",
        "risks_count": 5,
        "contracts_count": 12,
        "chart_data": [{"name": "Пн", "Тендеры": 10, "Выиграно": 2}],
        "tasks": [{"id": "1", "title": "Запустить парсер", "time": "Сейчас", "type": "info"}],
        "is_demo": False
    }

@app.post("/api/ai/export-risks-excel")
async def api_export_risks_excel(data: dict = Body(...)):
    """Экспорт результатов анализа рисков в Excel .xlsx"""
    logger.info("Excel export started")
    results = data.get('results', [])
    if not results:
        raise HTTPException(status_code=400, detail="No results to export")

    try:
        wb = openpyxl.Workbook()
        
        # 1. Сводка
        ws_summary = wb.active
        ws_summary.title = "Сводка"
        ws_summary.append(["ID Тендера", "Кол-во рисков", "Наличие противоречий", "Подробный отчет", "Ключевой вывод"])
        
        # 2. Краткие риски
        ws_risks = wb.create_sheet(title="Краткие риски")
        ws_risks.append(["ID Тендера", "Блок", "Что найдено", "Риск", "Что делать поставщику", "Источник", "Основание"])
        
        # 3. Подробный отчет
        ws_report = wb.create_sheet(title="Подробный отчет")
        ws_report.append(["ID Тендера", "Раздел", "Данные 1", "Данные 2", "Данные 3", "Данные 4", "Данные 5", "Данные 6"])
        
        # 4. Документы заявки
        ws_app_docs = wb.create_sheet(title="Документы заявки")
        ws_app_docs.append(["ID Тендера", "Документы в составе заявки", "Данные 2", "Данные 3", "Данные 4"])
        
        # 5. Документы при поставке
        ws_del_docs = wb.create_sheet(title="Документы при поставке")
        ws_del_docs.append(["ID Тендера", "Документы при поставке", "Данные 2", "Данные 3", "Данные 4"])
        
        # 6. Противоречия и соответствие
        ws_comp = wb.create_sheet(title="Противоречия и соответствие")
        ws_comp.append(["ID Тендера", "Раздел / Тип", "Описание", "Детали", "Основание"])
        
        # 7. Полный текст отчета
        ws_full_report = wb.create_sheet(title="Полный текст отчета")
        ws_full_report.append(["ID Тендера", "Полный текст отчета (Markdown)"])
        
        # 8. Источники и служебная информация
        ws_meta = wb.create_sheet(title="Источники и служебная инфо")
        ws_meta.append(["ID Тендера", "Тип инфо", "Содержание", "Детали"])
        
        # Style headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        for ws in wb.worksheets:
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            # Set default column width
            for i in range(1, ws.max_column + 1):
                ws.column_dimensions[get_column_letter(i)].width = 30
                
        for tender in results:
            tid = str(tender.get('id', 'N/A'))
            desc = tender.get('description', 'Нет описания')
            has_contract = "Да" if tender.get('has_contract') else "Нет"
            
            file_statuses = tender.get('file_statuses') or []
            if not isinstance(file_statuses, list): file_statuses = []
            file_count = len(file_statuses)
            
            rows = tender.get('rows') or []
            if not isinstance(rows, list): rows = []
            
            summary_notes = tender.get('summary_notes') or []
            if not isinstance(summary_notes, list): summary_notes = [str(summary_notes)]
            notes_full = "\n".join(summary_notes)
            desc_short = desc if len(desc) <= 150 else desc[:147] + "..."
            
            final_report_sections = tender.get('final_report_sections') or []
            if not isinstance(final_report_sections, (list, dict)): final_report_sections = []
            
            detailed_report = tender.get('detailed_report') or {}
            if not isinstance(detailed_report, dict): detailed_report = {}
            
            contradictions = tender.get('contradictions') or []
            if not isinstance(contradictions, list): contradictions = []
            
            final_report_markdown = tender.get('final_report_markdown') or ""
            if not isinstance(final_report_markdown, str): final_report_markdown = ""
            
            merged_sections = []
            
            logger.info(f"--- [EXCEL EXPORT PREPARATION FOR TENDER: {tid}] ---")
            logger.info(f"Rows count: {risk_count}")
            logger.info(f"Sections count: {len(final_report_sections)}")
            logger.info(f"Markdown length: {len(final_report_markdown)}")
            logger.info(f"Contradictions count: {len(contradictions)}")
            
            logger.info("--- [STRUCTURE QUALITY LOGGING] ---")
            if isinstance(final_report_sections, list):
                for sec in final_report_sections:
                    title = sec.get('section_title', 'Unknown')
                    s_data = sec.get('structured_data', [])
                    sub_sec = sec.get('sub_sections', {})
                    
                    if isinstance(s_data, list):
                        logger.info(f"Section '{title}': {len(s_data)} objects in structured_data")
                        if "3)" in title or "Проверка соответствия" in title:
                            logger.info(f" -> compliance_check explicitly has {len(s_data)} objects")
                    elif isinstance(s_data, dict):
                        logger.info(f"Section '{title}': {len(s_data.keys())} keys in structured_data (dict)")
                        if "3)" in title or "Проверка соответствия" in title:
                            logger.info(f" -> compliance_check explicitly has {len(s_data.keys())} keys")
                    elif isinstance(s_data, str):
                        logger.info(f"Section '{title}': 1 object (string)")
                        if "3)" in title or "Проверка соответствия" in title:
                            logger.info(f" -> compliance_check explicitly has 1 string object")
                    
                    if sub_sec:
                        in_app = sub_sec.get('in_application', [])
                        on_del = sub_sec.get('on_delivery', [])
                        in_app_count = len(in_app) if isinstance(in_app, list) else (1 if in_app else 0)
                        on_del_count = len(on_del) if isinstance(on_del, list) else (1 if on_del else 0)
                        logger.info(f"Section '{title}' -> documents_list.in_application: {in_app_count} objects")
                        logger.info(f"Section '{title}' -> documents_list.on_delivery: {on_del_count} objects")
            elif isinstance(final_report_sections, dict):
                for k, v in final_report_sections.items():
                    if isinstance(v, list):
                        logger.info(f"Section '{k}': {len(v)} objects")
                        if k == "compliance_check":
                            logger.info(f" -> compliance_check explicitly has {len(v)} objects")
                    elif isinstance(v, dict):
                        if k == "documents_list":
                            in_app = v.get('in_application', [])
                            on_del = v.get('on_delivery', [])
                            in_app_count = len(in_app) if isinstance(in_app, list) else (1 if in_app else 0)
                            on_del_count = len(on_del) if isinstance(on_del, list) else (1 if on_del else 0)
                            logger.info(f"Section '{k}' -> documents_list.in_application: {in_app_count} objects")
                            logger.info(f"Section '{k}' -> documents_list.on_delivery: {on_del_count} objects")
                        else:
                            logger.info(f"Section '{k}': {len(v.keys())} keys (dict)")
                            if k == "compliance_check":
                                logger.info(f" -> compliance_check explicitly has {len(v.keys())} keys")
                    elif isinstance(v, str):
                        logger.info(f"Section '{k}': 1 object (string)")
                        if k == "compliance_check":
                            logger.info(f" -> compliance_check explicitly has 1 string object")
            else:
                logger.info("final_report_sections is empty or not a list/dict.")
            logger.info("-----------------------------------")
            
            # 1. Populate parsed_sections from final_report_sections (Priority 1)
            parsed_sections = {}
            structured_sections = {}
            section_sources = {}
            sub_sections_data = {}

            if isinstance(final_report_sections, list):
                for sec in final_report_sections:
                    title = sec.get('section_title', '')
                    content = sec.get('content', '')
                    sub_sections = sec.get('sub_sections')
                    structured_data = sec.get('structured_data')
                    for i in range(1, 10):
                        if i not in parsed_sections and (f"Раздел {i}" in title or f"{i})" in title or f"{i}." in title):
                            parsed_sections[i] = content
                            section_sources[i] = "final_report_sections"
                            if sub_sections:
                                sub_sections_data[i] = sub_sections
                            if structured_data:
                                structured_sections[i] = structured_data
            elif isinstance(final_report_sections, dict):
                for k, v in final_report_sections.items():
                    content = "\n".join(v) if isinstance(v, list) else str(v)
                    for i in range(1, 10):
                        if i not in parsed_sections and (f"Раздел {i}" in k or f"{i})" in k or f"{i}." in k):
                            parsed_sections[i] = content
                            structured_sections[i] = v if isinstance(v, list) else []
                            section_sources[i] = "final_report_sections"

            # 2. Fallback to final_report_markdown parsing (Priority 2)
            # Removed as per user request: "Полностью перестать строить листы через повторный разбор markdown"
            
            if 1 in structured_sections and isinstance(structured_sections[1], list) and 2 in structured_sections and isinstance(structured_sections[2], list):
                risk_count = len(structured_sections[1]) + len(structured_sections[2])
            elif 1 in structured_sections and isinstance(structured_sections[1], list):
                risk_count = len(structured_sections[1])
            elif 2 in structured_sections and isinstance(structured_sections[2], list):
                risk_count = len(structured_sections[2])
            else:
                risk_count = len(rows)

            FALLBACK_TEXT = "Подробный отчет отсутствует в результате анализа"
            EMPTY_SECTION_TEXT = "Информация в предоставленной документации не обнаружена."

            def get_sec(num):
                content = parsed_sections.get(num, "")
                if not content:
                    section_sources[num] = "fallback"
                    return EMPTY_SECTION_TEXT if final_report_markdown else FALLBACK_TEXT
                return content

            compliance_content = get_sec(3)
            docs_content = get_sec(7)
            sub_docs = sub_sections_data.get(7, {})
            
            app_docs_content = ""
            del_docs_content = ""
            
            if sub_docs:
                in_app = sub_docs.get("in_application", [])
                on_del = sub_docs.get("on_delivery", [])
                app_docs_content = "\n".join([f"- {i}" for i in in_app]) if in_app else ""
                del_docs_content = "\n".join([f"- {i}" for i in on_del]) if on_del else ""
            
            if not app_docs_content and not del_docs_content:
                app_docs_content = docs_content
                del_docs_content = docs_content
            
            if not app_docs_content:
                app_docs_content = EMPTY_SECTION_TEXT if final_report_markdown else FALLBACK_TEXT
            if not del_docs_content:
                del_docs_content = EMPTY_SECTION_TEXT if final_report_markdown else FALLBACK_TEXT

            if 3 in structured_sections and isinstance(structured_sections[3], list):
                has_contradictions = "Да" if len(structured_sections[3]) > 0 else "Нет"
            else:
                has_contradictions = "Да" if compliance_content and compliance_content not in (FALLBACK_TEXT, EMPTY_SECTION_TEXT) and ("противореч" in compliance_content.lower() or "ошибк" in compliance_content.lower()) and "не обнаружено" not in compliance_content.lower() else "Нет"
            
            # --- Visual separation for all sheets except Сводка ---
            for ws in [ws_risks, ws_report, ws_app_docs, ws_del_docs, ws_comp, ws_full_report, ws_meta]:
                ws.append([f"ТЕНДЕР: {tid}"])
                ws[ws.max_row][0].font = Font(bold=True, size=12, color="FFFFFF")
                ws[ws.max_row][0].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=ws.max_column if ws.max_column > 1 else 2)
                
                ws.append([f"Описание: {desc}"])
                ws[ws.max_row][0].font = Font(italic=True)
                ws[ws.max_row][0].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=ws.max_column if ws.max_column > 1 else 2)
            
            # --- Build Подробный отчет ---
            # We want to show all 9 sections here
            section_names = [
                "1) Риски участия и исполнения договора",
                "2) Риски недопуска заявки и потери баллов",
                "3) Проверка соответствия документации и закона",
                "4) Условия поставки и приемки",
                "5) Условия оплаты",
                "6) Ответственность сторон",
                "7) Перечень документов",
                "8) Требования по реестрам и ограничениям",
                "9) Рекомендации Поставщику"
            ]
            
            for i, s_name in enumerate(section_names, 1):
                first_row_in_sec = True
                
                if i in structured_sections and structured_sections[i]:
                    sec_data = structured_sections[i]
                    if isinstance(sec_data, list) and len(sec_data) > 0:
                        for item in sec_data:
                            if isinstance(item, dict):
                                row_data = [clean_markdown(str(v)) for v in item.values()]
                                ws_report.append([tid, s_name if first_row_in_sec else ""] + row_data)
                                first_row_in_sec = False
                            elif isinstance(item, str):
                                ws_report.append([tid, s_name if first_row_in_sec else "", clean_markdown(item)])
                                first_row_in_sec = False
                    elif isinstance(sec_data, dict) and len(sec_data) > 0:
                        for k, v in sec_data.items():
                            ws_report.append([tid, s_name if first_row_in_sec else "", clean_markdown(str(k)), clean_markdown(str(v))])
                            first_row_in_sec = False
                    elif isinstance(sec_data, str) and len(sec_data) > 0:
                        ws_report.append([tid, s_name if first_row_in_sec else "", clean_markdown(sec_data)])
                        first_row_in_sec = False
                    else:
                        ws_report.append([tid, s_name, EMPTY_SECTION_TEXT])
                elif i in sub_sections_data and sub_sections_data[i]:
                    sec_data = sub_sections_data[i]
                    if isinstance(sec_data, dict) and len(sec_data) > 0:
                        # Handle section 7 (documents_list) which is a dict
                        for sub_k, sub_v in sec_data.items():
                            sub_title = "В составе заявки" if sub_k == "in_application" else "При поставке"
                            ws_report.append([tid, s_name if first_row_in_sec else "", sub_title])
                            first_row_in_sec = False
                            if isinstance(sub_v, list):
                                for item in sub_v:
                                    if isinstance(item, dict):
                                        row_data = [clean_markdown(str(v)) for v in item.values()]
                                        ws_report.append([tid, "", ""] + row_data)
                                    elif isinstance(item, str):
                                        ws_report.append([tid, "", "", clean_markdown(item)])
                            elif isinstance(sub_v, str):
                                ws_report.append([tid, "", "", clean_markdown(sub_v)])
                    else:
                        ws_report.append([tid, s_name, EMPTY_SECTION_TEXT])
                else:
                    ws_report.append([tid, s_name, EMPTY_SECTION_TEXT])
            
            # --- Build Полный текст отчета ---
            if final_report_markdown:
                ws_full_report.append([tid, final_report_markdown])
            else:
                ws_full_report.append([tid, FALLBACK_TEXT])
            
            has_report = "Да" if final_report_markdown or parsed_sections else "Нет"
            ws_summary.append([tid, risk_count, has_contradictions, has_report, notes_full])
            
            # --- Build Документы заявки ---
            if 7 in sub_sections_data and isinstance(sub_sections_data[7], dict) and "in_application" in sub_sections_data[7]:
                in_app = sub_sections_data[7]["in_application"]
                if in_app:
                    if isinstance(in_app, list):
                        for doc_item in in_app:
                            if isinstance(doc_item, dict):
                                ws_app_docs.append([tid] + [clean_markdown(str(v)) for v in doc_item.values()])
                            elif isinstance(doc_item, str):
                                ws_app_docs.append([tid, clean_markdown(doc_item)])
                    elif isinstance(in_app, str):
                        ws_app_docs.append([tid, clean_markdown(in_app)])
                else:
                    ws_app_docs.append([tid, EMPTY_SECTION_TEXT])
            else:
                ws_app_docs.append([tid, EMPTY_SECTION_TEXT])

            # --- Build Документы при поставке ---
            if 7 in sub_sections_data and isinstance(sub_sections_data[7], dict) and "on_delivery" in sub_sections_data[7]:
                on_del = sub_sections_data[7]["on_delivery"]
                if on_del:
                    if isinstance(on_del, list):
                        for doc_item in on_del:
                            if isinstance(doc_item, dict):
                                ws_del_docs.append([tid] + [clean_markdown(str(v)) for v in doc_item.values()])
                            elif isinstance(doc_item, str):
                                ws_del_docs.append([tid, clean_markdown(doc_item)])
                    elif isinstance(on_del, str):
                        ws_del_docs.append([tid, clean_markdown(on_del)])
                else:
                    ws_del_docs.append([tid, EMPTY_SECTION_TEXT])
            else:
                ws_del_docs.append([tid, EMPTY_SECTION_TEXT])

            # --- Build Противоречия и соответствие ---
            if 3 in structured_sections and isinstance(structured_sections[3], list) and len(structured_sections[3]) > 0:
                for item in structured_sections[3]:
                    if isinstance(item, dict):
                        ws_comp.append([tid] + [clean_markdown(str(v)) for v in item.values()])
                    elif isinstance(item, str):
                        ws_comp.append([tid, clean_markdown(item)])
            elif 3 in structured_sections and isinstance(structured_sections[3], dict) and len(structured_sections[3]) > 0:
                for k, v in structured_sections[3].items():
                    ws_comp.append([tid, clean_markdown(str(k)), clean_markdown(str(v))])
            elif 3 in structured_sections and isinstance(structured_sections[3], str) and len(structured_sections[3]) > 0:
                ws_comp.append([tid, clean_markdown(structured_sections[3])])
            else:
                ws_comp.append([tid, EMPTY_SECTION_TEXT])
            
            logger.info(f"Excel export for {tid}: {risk_count} rows, {len(parsed_sections)} sections, markdown length: {len(final_report_markdown)}, contradictions: {has_contradictions}")
            logger.info(f"Sections found for Excel: {list(parsed_sections.keys())}")
            
            logger.info("--- [EXCEL DATA SOURCES SUMMARY] ---")
            logger.info("Краткие риски -> rows")
            logger.info("Подробный отчет -> final_report_sections")
            logger.info("Документы заявки -> final_report_sections.documents_list.in_application")
            logger.info("Документы при поставке -> final_report_sections.documents_list.on_delivery")
            logger.info("Противоречия и соответствие -> final_report_sections.compliance_check")
            logger.info("Полный текст отчета -> final_report_markdown")
            logger.info("------------------------------------")
            
            # --- Build Краткие риски ---
            risks_added = False
            for row in rows:
                if not isinstance(row, dict):
                    continue
                ws_risks.append([
                    tid,
                    clean_markdown(row.get('block', '')),
                    clean_markdown(row.get('finding', '')),
                    clean_markdown(row.get('risk_level', '')),
                    clean_markdown(row.get('supplier_action', '')),
                    clean_markdown(f"{row.get('source_document', '')} {row.get('source_reference', '')}".strip()),
                    clean_markdown(row.get('legal_basis', ''))
                ])
                risks_added = True
                
            if not risks_added:
                ws_risks.append([tid, "Риски отсутствуют или не найдены", "-", "-", "-", "-", "-"])
            
            # --- Build Источники и служебная инфо ---
            # 1. File statuses
            file_info = "\n".join([f"{f.get('filename', 'N/A')}: {f.get('status', 'N/A')}" for f in file_statuses if isinstance(f, dict)])
            ws_meta.append([tid, "Статусы файлов", file_info, ""])
            
            # 2. Classification notes
            class_notes = "\n".join(tender.get('classification_notes', []) if isinstance(tender.get('classification_notes'), list) else [])
            if class_notes:
                ws_meta.append([tid, "Заметки классификации", class_notes, ""])
                
            # 3. Sources from markdown and structured_sections
            md_sources = set()
            if final_report_markdown:
                bracket_matches = re.findall(r'\[(.*?)\]', final_report_markdown)
                for m in bracket_matches:
                    if any(kw in m.lower() for kw in ['договор', 'контракт', 'тз', 'задан', 'п.', 'раздел', 'ст.', 'фз', 'закон', 'документаци', 'приложени', 'часть']):
                        md_sources.add(m.strip())
                source_matches = re.findall(r'(?i)источник[и]?\s*[:\-]?\s*([^\n\.\;]+)', final_report_markdown)
                for m in source_matches:
                    md_sources.add(m.strip())
                    
            for sec_num, sec_data in structured_sections.items():
                if isinstance(sec_data, list):
                    for item in sec_data:
                        if isinstance(item, dict) and 'source' in item and item['source']:
                            md_sources.add(str(item['source']).strip())
                elif isinstance(sec_data, dict):
                    for k, v in sec_data.items():
                        if isinstance(v, list):
                            for item in v:
                                if isinstance(item, dict) and 'source' in item and item['source']:
                                    md_sources.add(str(item['source']).strip())
            
            for src in md_sources:
                ws_meta.append([tid, "Источник (из отчета)", clean_markdown(src), ""])
            
            # 4. Sources from rows (if unique)
            row_sources = {f"{r.get('source_document', '')} {r.get('source_reference', '')}".strip() for r in rows if isinstance(r, dict)}
            for src in row_sources:
                if src and src not in md_sources:
                    ws_meta.append([tid, "Источник (из рисков)", clean_markdown(src), ""])
                
            logger.info(f"--- [EXCEL EXPORT COMPLETED FOR TENDER: {tid}] ---")
            logger.info(f"Lengths -> final_report_markdown: {len(final_report_markdown)}, app_docs: {len(app_docs_content)}, del_docs: {len(del_docs_content)}, compliance: {len(compliance_content)}")
            logger.info(f"Counts -> final_report_sections: {len(final_report_sections)}, rows: {risk_count}, contradictions: {len(contradictions)}")
            
            logger.info(f"--- [EXCEL SHEET SOURCES FOR TENDER: {tid}] ---")
            logger.info(f"Sheet 'Краткие риски': rows")
            logger.info(f"Sheet 'Подробный отчет': { {f'Раздел {k}': v for k, v in section_sources.items()} }")
            logger.info(f"Sheet 'Документы заявки': {section_sources.get(7, 'fallback')}")
            logger.info(f"Sheet 'Документы при поставке': {section_sources.get(7, 'fallback')}")
            logger.info(f"Sheet 'Противоречия и соответствие': {section_sources.get(3, 'fallback')}")
            logger.info(f"Sheet 'Полный текст отчета': {'final_report_markdown' if final_report_markdown else 'fallback'}")
            logger.info(f"Sheet 'Источники и служебная инфо': rows + final_report_markdown")
            
            # Empty row before next tender
            for ws in [ws_risks, ws_report, ws_app_docs, ws_del_docs, ws_comp, ws_full_report, ws_meta]:
                ws.append([])

        logger.info("Excel export data prepared successfully")

        # Add filters
        # Only apply auto_filter to the first row to avoid issues with merged cells
        ws_risks.auto_filter.ref = f"A1:{get_column_letter(ws_risks.max_column)}1"
        ws_report.auto_filter.ref = f"A1:{get_column_letter(ws_report.max_column)}1"
        ws_app_docs.auto_filter.ref = f"A1:{get_column_letter(ws_app_docs.max_column)}1"
        ws_del_docs.auto_filter.ref = f"A1:{get_column_letter(ws_del_docs.max_column)}1"
        ws_comp.auto_filter.ref = f"A1:{get_column_letter(ws_comp.max_column)}1"
        ws_full_report.auto_filter.ref = f"A1:{get_column_letter(ws_full_report.max_column)}1"
        ws_meta.auto_filter.ref = f"A1:{get_column_letter(ws_meta.max_column)}1"

        # Set column widths and wrap text
        for ws in wb.worksheets:
            for col_idx in range(1, ws.max_column + 1):
                column = get_column_letter(col_idx)
                if ws.title == "Сводка":
                    widths = {'A': 15, 'B': 15, 'C': 20, 'D': 20, 'E': 60}
                    ws.column_dimensions[column].width = widths.get(column, 20)
                elif ws.title == "Краткие риски":
                    widths = {'A': 15, 'B': 20, 'C': 40, 'D': 15, 'E': 40, 'F': 30, 'G': 30}
                    ws.column_dimensions[column].width = widths.get(column, 20)
                elif ws.title == "Подробный отчет":
                    widths = {'A': 15, 'B': 40, 'C': 100}
                    ws.column_dimensions[column].width = widths.get(column, 20)
                elif ws.title in ["Документы заявки", "Документы при поставке", "Противоречия и соответствие"]:
                    widths = {'A': 15, 'B': 100}
                    ws.column_dimensions[column].width = widths.get(column, 20)
                elif ws.title == "Полный текст отчета":
                    widths = {'A': 15, 'B': 120}
                    ws.column_dimensions[column].width = widths.get(column, 20)
                elif ws.title == "Источники и служебная инфо":
                    widths = {'A': 15, 'B': 30, 'C': 80, 'D': 40}
                    ws.column_dimensions[column].width = widths.get(column, 20)
                
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    if type(cell).__name__ == 'MergedCell':
                        continue
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                    cell.border = thin_border

        # Save to temp file
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            wb.save(tmp.name)
            logger.info("Excel export finished successfully")
            return FileResponse(tmp.name, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", filename="tender_risks_report.xlsx")

    except Exception as e:
        logger.error(f"Excel Export Error: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))

if __name__ == "__main__":
    uvicorn.run("backend.main:app", host="0.0.0.0", port=8000, reload=True)
